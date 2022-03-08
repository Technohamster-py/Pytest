"""
--------------------------------------------------
Programm by: Technohamster

Python version:        Year:
        3.8            2020

Thank you for using my programm! Good luck)
---------------------------------------------------
"""

import openpyxl, converter


def write(path, students_list, task_number):
    """

    :param path: Путь к таблице
    :param students_list: имена с оценками
    :param task_number: номер задания (task_number + 1 = номер столбца)
    :return: None
    """
    book = openpyxl.load_workbook(path, data_only=True)
    ws = book.active
    log = ''
    errors = ''

    names_matrix, row_number = get_matrix(ws)
    task_column = converter.convert(task_number + 1)
    log += 'Начинается запись оценок в таблицу\n'
    for name in students_list.keys():
        if name in names_matrix.keys():
            cell = f'{task_column}{names_matrix[name]}'
            ws[cell] = students_list[name]
        else:
            row_number += 1
            cell = f'{task_column}{row_number}'
            ws[cell] = students_list[name]
            errors += f'В таблице не существует имя {name}, имя было добавлено в таблицу'
    log += 'Запись оценок в таблицу завершена'
    return log, errors


def get_matrix(sheet):
    """
    Считывание имен учеников из первого столбца таблицы
    :param sheet: активный лист страницы
    :return: имена, последняя заполненная трочка в таблице
    """
    row = 1
    names = {}
    while True:
        name = sheet.cell(row=row,  column=1).value
        if name == '':
            end_row = row - 1
            break
        else:
            names[name] = row

    return names, end_row
